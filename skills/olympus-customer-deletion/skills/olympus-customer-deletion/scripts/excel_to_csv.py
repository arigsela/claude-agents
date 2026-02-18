#!/usr/bin/env python3
"""
Convert Olympus customer termination Excel attachments to the CSV format
required by delete_with_ai.py.

Output format: file_path,company,date,notes

The Olympus termination Excel has two sheets:
  - "Customer Termination Info": metadata (client name, customer key)
  - "Complete File List": actual file list with FILE PATH + FILE NAME columns

Usage:
    python3 excel_to_csv.py \
        --input "tickets/DEVOPS-8035/Terminated Clients - Atlantic Constructors.xlsx" \
        --output ".ralph-projects/olympus-customer-deletions/deletion_lists/DEVOPS-8035-atlantic-constructors.csv"

The script automatically:
  - Reads the "Complete File List" sheet (or first sheet with data)
  - Combines FILE PATH + FILE NAME columns into the full S3 key
  - Strips the S3 bucket name prefix (e.g. "artemis-client-etl-data/")
  - Extracts company and date from the path
"""

import argparse
import csv
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# S3 bucket name prefix to strip from paths
S3_BUCKET_PREFIX = "artemis-client-etl-data/"

# Sheet preference order
PREFERRED_SHEETS = ["Complete File List", "File List", "Files", "Deletion List"]

# Path regex to extract company and date
# automation_drop/prod/{company}/{user}/{date}/{file}
PATH_RE = re.compile(
    r"automation_drop/(?:prod|dev)/([^/]+)/[^/]+/(\d{4}-\d{2}(?:-\d{2})?)/.*"
)


def find_sheet(wb):
    """Return the best sheet for file list data."""
    for name in PREFERRED_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    # Fall back to first sheet with data
    return wb.worksheets[0]


def find_header_row(rows):
    """Return (row_index, col_indices) for the header row containing FILE PATH."""
    for i, row in enumerate(rows[:20]):
        row_lower = [str(c).strip().lower() if c else "" for c in row]
        if any("file path" in c or "filepath" in c or "path" in c for c in row_lower):
            # Map column names to indices
            return i, {v: idx for idx, v in enumerate(row_lower) if v}
    return None, None


def normalize_path(file_path, file_name):
    """Combine path + filename, strip bucket prefix, normalize slashes."""
    if file_path is None:
        file_path = ""
    if file_name is None:
        file_name = ""

    file_path = str(file_path).strip().rstrip("/")
    file_name = str(file_name).strip()

    # Combine path and filename
    if file_path and file_name:
        full = file_path + "/" + file_name
    elif file_path:
        full = file_path
    else:
        full = file_name

    # Strip bucket prefix
    if full.startswith(S3_BUCKET_PREFIX):
        full = full[len(S3_BUCKET_PREFIX):]

    # Normalize double slashes
    full = re.sub(r"/+", "/", full).strip("/")

    return full


def extract_company_date(path):
    """Extract company and date from automation_drop path."""
    m = PATH_RE.match(path)
    if m:
        return m.group(1), m.group(2)
    return "", ""


def main():
    parser = argparse.ArgumentParser(description="Convert customer termination Excel to CSV")
    parser.add_argument("--input", required=True, help="Path to the Excel (.xlsx) file")
    parser.add_argument("--output", required=True, help="Path for the output CSV file")
    parser.add_argument("--sheet", default=None, help="Sheet name to use (default: auto-detect)")
    parser.add_argument("--default-company", default="", help="Default company if not detectable from path")
    parser.add_argument("--default-notes", default="Terminated", help="Default notes value")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: Excel file not found: {args.input}")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    print(f"Excel file:  {args.input}")
    print(f"Sheets:      {wb.sheetnames}")

    # Select sheet
    if args.sheet:
        if args.sheet in wb.sheetnames:
            ws = wb[args.sheet]
        else:
            try:
                ws = wb.worksheets[int(args.sheet)]
            except (ValueError, IndexError):
                print(f"ERROR: Sheet '{args.sheet}' not found.")
                sys.exit(1)
    else:
        ws = find_sheet(wb)

    print(f"Using sheet: {ws.title}")

    rows = list(ws.iter_rows(values_only=True))
    print(f"Total rows:  {len(rows)}")

    # Find header row
    header_idx, col_map = find_header_row(rows)

    if header_idx is None:
        print("\nERROR: Could not find header row with FILE PATH column.")
        print("First 5 rows of data:")
        for i, row in enumerate(rows[:5]):
            print(f"  Row {i+1}: {row}")
        print("\nRe-run with --sheet <sheet_name> to specify a different sheet.")
        sys.exit(1)

    print(f"Header row:  {header_idx + 1}")
    print(f"Columns:     {list(col_map.keys())}")

    # Find path and filename column indices (exact match preferred, avoid NOTE columns)
    path_col_idx = None
    name_col_idx = None

    for col_name, idx in col_map.items():
        # Exact matches for common path column names (not "file name")
        if col_name in ("file path", "filepath", "path", "s3_path", "s3 path"):
            path_col_idx = idx
        # Exact matches for filename column names
        elif col_name in ("file name", "filename", "file"):
            name_col_idx = idx

    if path_col_idx is None:
        # Try to find a column that has automation_drop in first data row
        data_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else None
        if data_row:
            for idx, val in enumerate(data_row):
                if val and "automation_drop" in str(val):
                    path_col_idx = idx
                    print(f"Auto-detected path column at index {idx}")
                    break

    if path_col_idx is None:
        print(f"\nERROR: Could not determine file path column from: {list(col_map.keys())}")
        sys.exit(1)

    print(f"\nMapping:")
    print(f"  file_path <- column {path_col_idx} (FILE PATH) + column {name_col_idx} (FILE NAME)")

    # Create output directory if needed
    out_dir = os.path.dirname(os.path.abspath(args.output))
    os.makedirs(out_dir, exist_ok=True)

    written = 0
    skipped = 0

    with open(args.output, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["file_path", "company", "date", "notes"])

        for row in rows[header_idx + 1:]:
            # Get path parts
            path_val = row[path_col_idx] if path_col_idx < len(row) else None
            name_val = row[name_col_idx] if name_col_idx is not None and name_col_idx < len(row) else None

            file_path = normalize_path(path_val, name_val)

            if not file_path:
                skipped += 1
                continue

            company, date = extract_company_date(file_path)
            if not company:
                company = args.default_company

            writer.writerow([file_path, company, date, args.default_notes])
            written += 1

    print(f"\nOutput CSV:  {args.output}")
    print(f"  Written:   {written} rows")
    print(f"  Skipped:   {skipped} empty rows")

    if written == 0:
        print("\nWARNING: No rows were written. The Excel structure may be unexpected.")
        print("First 5 data rows:")
        for row in rows[header_idx + 1: header_idx + 6]:
            print(f"  {row}")
        sys.exit(1)

    # Preview
    print(f"\nPreview (first 10 rows of {args.output}):")
    print(f"{'file_path':<70} {'company':<12} {'date':<10} notes")
    print("-" * 110)
    with open(args.output, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            if i >= 10:
                print(f"  ... ({written - 10} more rows)")
                break
            fp = row["file_path"]
            fp_display = fp[:67] + "..." if len(fp) > 70 else fp
            print(f"{fp_display:<70} {row['company']:<12} {row['date']:<10} {row['notes']}")

    print(f"\nDone. Review preview above, then proceed to Step 3 (dry-run).")


if __name__ == "__main__":
    main()
